"""
ingestion/build_excel_index.py
Extracts structured metadata from all xlsx files in data/excel/
Merges with PDF index to create the unified metadata store.
Saves to index/unified_metadata_store.pkl

Run from project root:
  python3 src/ingestion/build_excel_index.py
  python3 src/ingestion/build_excel_index.py --rebuild
"""

import sys, os
sys.path.insert(0, os.path.join(os.path.dirname(__file__), ".."))

import pickle, re, hashlib
import numpy as np
import openpyxl
from pathlib import Path
from datetime import datetime
from sklearn.feature_extraction.text import TfidfVectorizer
from config import EXCEL_DIR, PDF_INDEX, UNIFIED_STORE, INDEX_DIR, ensure_dirs

# ── ABBREVIATION DICTIONARY ───────────────────────────────────────────────────
# Maps cryptic column abbreviations to plain English.
# Covers Honeywell BA domain: equipment, energy, commissioning, sensors, financial.

ABBREV_MAP = {
    "bld":"building","fac":"facility","flr":"floor","zn":"zone","rm":"room",
    "bld_ref":"building reference","fac_cd":"facility code","rec_id":"record ID",
    "prd_key":"period key","proj_no":"project number","cx_id":"commissioning ID",
    "dev_cat":"device category","eq":"equipment","eq_cat":"equipment category",
    "eq_subcat":"equipment subcategory",
    "mfr":"manufacturer","mfr_cd":"manufacturer code","mdl":"model","mdl_cd":"model code",
    "srl_no":"serial number","fw_ver":"firmware version","hw_rev":"hardware revision",
    "cap_val":"capacity value","cap_unt":"capacity unit","prot_cls":"protection class",
    "inst_dt":"installation date","comm_dt":"commissioning date","wrnt_exp":"warranty expiry",
    "lst_pm_dt":"last preventive maintenance date","nxt_pm_due":"next PM due date",
    "lst_cm_dt":"last corrective maintenance date","cx_dt":"commissioning date",
    "retst_dt":"retest date","cert_dt":"certificate date","cert_exp":"certificate expiry",
    "handover_dt":"handover date",
    "pm":"preventive maintenance","cm":"corrective maintenance",
    "svc_prvdr":"service provider","wo_ref":"work order reference",
    "pm_intv_mo":"PM interval months","cm_hrs":"corrective maintenance hours",
    "cm_cst":"corrective maintenance cost","flt_cnt":"fault count",
    "dwntime_hr":"downtime hours","flt_cnt_12m":"fault count last 12 months",
    "cur_tmp_c":"current temperature Celsius","cur_hum_pct":"current humidity percent",
    "cur_co2_ppm":"current CO2 ppm","cur_pwr_kw":"current power kW",
    "alm_cnt":"alarm count","alm_cnt_7d":"alarm count last 7 days",
    "alm_cnt_30d":"alarm count last 30 days","alm_30d_tot":"total alarms last 30 days",
    "avail_pct":"availability percent","resp_ms":"response time ms",
    "lst_rdg_ts":"last reading timestamp","rdg_qual":"reading quality",
    "purch_cst":"purchase cost","bk_val":"book value","bk_val_cur":"current book value",
    "dep_rt":"depreciation rate","dep_mth":"depreciation method",
    "capex":"capital expenditure","opex_ann":"annual operational expenditure",
    "lcoe_calc":"levelised cost of energy",
    "el_kwh":"electricity kWh","kwh":"kilowatt hours",
    "el_kwh_t":"total electricity kWh","el_kwh_hvac":"HVAC electricity kWh",
    "el_kwh_ltg":"lighting electricity kWh","el_kwh_it":"IT electricity kWh",
    "el_pkdmd_kw":"peak demand kW","el_pf":"power factor","el_thd_pct":"total harmonic distortion percent",
    "gs_m3_t":"total gas cubic metres","gs_kwh_eq":"gas kWh equivalent",
    "wt_m3_t":"total water cubic metres","wt_m3_cool":"cooling water cubic metres",
    "th_ctl_mwh":"thermal cooling MWh","th_htg_mwh":"thermal heating MWh",
    "chil_cop":"chiller COP","boil_eff":"boiler efficiency",
    "hvac_run_hr":"HVAC run hours","cop_cool":"COP cooling","cop_htg":"COP heating",
    "co2e":"CO2 equivalent","co2e_el_t":"CO2 equivalent electricity tonnes",
    "co2e_gs_t":"CO2 equivalent gas tonnes","co2e_t_tot":"total CO2 equivalent tonnes",
    "co2e_intns_kgsqm":"CO2 intensity kg per sq metre",
    "ghg":"greenhouse gas","ghg_scpe1":"GHG scope 1","ghg_scpe2":"GHG scope 2",
    "gresb_scr":"GRESB score","enstar_rtg":"Energy Star rating",
    "nrg_cost":"energy cost","nrg_bgt":"energy budget","bgt_var":"budget variance",
    "nrg_intns":"energy intensity","trgt_kwh":"target kWh","trgt_var":"target variance",
    "iso50001_cmp":"ISO 50001 compliance","leed_pts":"LEED points",
    "ip_addr":"IP address","mac_addr":"MAC address","net_seg":"network segment",
    "baud_rt":"baud rate","comm_prt":"communication protocol","poll_iv":"polling interval",
    "wbus_ad":"wallbus address","scl_grp":"scale group","bac_dev_id":"BACnet device ID",
    "bac_net_no":"BACnet network number","bac_obj_cnt":"BACnet object count",
    "slc_no":"SLC loop number","slc_ad":"SLC address",
    "sp_heat":"heating setpoint","sp_cool":"cooling setpoint",
    "sp_co2":"CO2 setpoint","sp_hum":"humidity setpoint",
    "sp_heat_c":"heating setpoint Celsius","sp_cool_c":"cooling setpoint Celsius",
    "sp_co2_ppm":"CO2 setpoint ppm","sp_hum_pct":"humidity setpoint percent",
    "do_qty":"digital output count","di_qty":"digital input count",
    "ao_qty":"analogue output count","ai_qty":"analogue input count",
    "crit_flg":"critical flag","insp_sts":"inspection status","doc_ref":"document reference",
    "cx_sts":"commissioning status","snag_no":"snag number","snag_desc":"snag description",
    "retst_sts":"retest status","cert_no":"certificate number","rpt_sts":"report status",
    "verif_by":"verified by","aud_flg":"audit flag","aud_by":"audited by",
    "tot_eq":"total equipment","act_eq":"active equipment","flt_eq":"faulty equipment",
    "pm_ovrd":"PM overdue","avg_avail":"average availability",
    "opex_ann_tot":"total annual OPEX","capex_yr_tot":"total CAPEX this year",
    "crit_cnt":"critical count",
}

DOMAIN_KEYWORDS = {
    "equipment_maintenance": ["eq_cat","pm","cm","svc","wo","flt","dwntime","mfr","srl","fw","inst","wrnt","alm","fault","alarm","event","count"],
    "energy_utilities":      ["kwh","co2","nrg","el_","gs_","wt_","cop","hvac","boil","chil","th_","ghg","gresb"],
    "commissioning_testing": ["cx_","cx_sts","snag","cert","retst","handover","comm_chk","db_chk","io_chk"],
    "sensor_realtime":       ["cur_tmp","cur_hum","cur_co2","cur_pwr","alm_cnt","rdg_qual","avail_pct"],
    "financial":             ["cst","cost","val","bgt","capex","opex","dep","lcoe","roi","purch"],
    "network_config":        ["ip_addr","mac","vlan","baud","proto","bac_","slc_","wbus"],
    "location_asset":        ["bld_ref","flr","zn","rec_id","fac_cd"],
}

DOMAIN_SUMMARIES = {
    "equipment_maintenance": (
        "Equipment and asset maintenance tracking. "
        "Records building automation devices: category, manufacturer, model, firmware, "
        "installation date, warranty. Maintenance history: PM/CM dates, work orders, "
        "service provider, costs. Performance: fault count, downtime, availability. "
        "Financial: purchase cost, book value, depreciation. "
        "Use to answer: equipment status, fault history, overdue maintenance, "
        "warranty coverage, operational costs."
    ),
    "energy_utilities": (
        "Monthly energy and utilities consumption data. "
        "Electricity (kWh), gas (m³), water (m³), thermal energy across multiple buildings. "
        "Includes: HVAC energy, lighting, IT load, peak demand, carbon emissions (CO2e), "
        "carbon intensity per sq.m., GHG scope 1/2/3, GRESB score, Energy Star rating. "
        "Financial: energy cost vs budget variance. Weather: HDD, CDD, avg temperature. "
        "Use to answer: energy trends, carbon footprint, budget variance, sustainability."
    ),
    "commissioning_testing": (
        "Building automation commissioning and test records. "
        "Device types: IQ4 controllers, TC300 thermostats, RS-WMB/RD-WMB room displays, "
        "AFP-3030 fire panels, FCM-1 modules, VAV/FCU controllers. "
        "Network config: IP, MAC, VLAN, BACnet device ID, SLC address, wallbus address. "
        "Test results: comms check, DB check, I/O check, alarm check, commissioning status. "
        "Snag descriptions, retest records, certificate numbers. "
        "Use to answer: commissioning status, test failures, open snags, certification."
    ),
    "sensor_realtime": (
        "Real-time sensor readings: temperature, CO2, humidity, occupancy, HVAC, alarms. "
        "Contains current and historical readings with alert flags. "
        "Use to answer: current building conditions, sensor alerts, performance trends."
    ),
    "financial": (
        "Financial and cost data: purchase costs, book values, depreciation, CAPEX, OPEX. "
        "Use to answer: asset valuation, cost analysis, budget planning."
    ),
    "network_config": (
        "Network and communications configuration: IP addresses, BACnet device IDs, "
        "wallbus addresses, SLC assignments, communication protocols. "
        "Use to answer: device network config, addressing, protocol settings."
    ),
    "location_asset": (
        "Asset location and identification data: buildings, floors, zones, asset IDs. "
        "Use to answer: where assets are located, building inventory."
    ),
    "general": (
        "Building automation operational data across multiple columns and buildings. "
        "Use to answer: general equipment and operational status queries."
    ),
}


def expand_col(col_name: str) -> str:
    """Map a cryptic column name to a human-readable meaning."""
    key = col_name.lower().split("::")[-1]
    if key in ABBREV_MAP:
        return ABBREV_MAP[key]
    # Longest-prefix match
    best = max((a for a in ABBREV_MAP if key.startswith(a)), key=len, default="")
    if best:
        suffix = key[len(best):].strip("_")
        return ABBREV_MAP[best] + (f" ({suffix})" if suffix else "")
    # Split on _ and expand each part
    return " ".join(ABBREV_MAP.get(p, p) for p in key.split("_"))


def profile_sheet(filepath: Path, sheet_name: str) -> dict:
    """Read only the first 12 rows to understand structure — never the full file."""
    try:
        wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
    except Exception as e:
        return {}
    ws = wb[sheet_name]
    all_rows   = list(ws.iter_rows(max_row=12, values_only=True))
    total_rows = ws.max_row or 0
    total_cols = ws.max_column or 0
    wb.close()

    if not all_rows:
        return {}

    def is_header(row):
        vals = [v for v in row if v is not None]
        return bool(vals) and sum(1 for v in vals if isinstance(v,str)) / len(vals) > 0.6

    header_rows, data_start = [], 0
    for i, row in enumerate(all_rows[:4]):
        if is_header(row):
            header_rows.append([str(v or "").strip() for v in row])
            data_start = i + 1
        else:
            break

    if not header_rows:
        return {}

    # Flatten merged headers: "SectionName::ColName"
    if len(header_rows) == 2:
        sec_row, col_row = header_rows[0], header_rows[1]
        flat, cur_sec = [], ""
        for s, c in zip(sec_row, col_row):
            if s: cur_sec = s
            flat.append(f"{cur_sec}::{c}" if cur_sec and c else (c or s or ""))
    else:
        flat = header_rows[0]

    sample_rows = [
        [str(v or "")[:50] for v in row[:len(flat)]]
        for row in all_rows[data_start:data_start+8]
        if any(v is not None for v in row)
    ]

    return {
        "sheet_name": sheet_name, "total_rows": total_rows, "total_cols": total_cols,
        "flat_headers": flat, "sample_rows": sample_rows[:3],
        "data_start_row": data_start + 1,
    }


def build_excel_record(filepath: Path, sheet_name: str) -> dict:
    profile = profile_sheet(filepath, sheet_name)
    if not profile:
        return {}

    flat_hdrs = profile.get("flat_headers", [])
    sem_map   = {col: expand_col(col) for col in flat_hdrs}

    header_str   = " ".join(h.lower() for h in flat_hdrs)
    domain_scores = {d: sum(1 for kw in kws if kw in header_str)
                     for d, kws in DOMAIN_KEYWORDS.items()}
    primary = max(domain_scores, key=domain_scores.get) if any(domain_scores.values()) else "general"

    # Detect sample locations
    locations = set()
    for row in profile.get("sample_rows",[]):
        for ci, val in enumerate(row):
            if ci < len(flat_hdrs) and val:
                col_l = flat_hdrs[ci].lower()
                if any(kw in col_l for kw in ["bld_ref","bld","fac","loc"]) and len(val) < 20:
                    locations.add(val)
    loc_str = ", ".join(sorted(locations)[:4]) or "multiple buildings"

    summary = (
        f"File: {filepath.name} | Sheet: {sheet_name} | "
        f"{profile['total_rows']} rows | {profile['total_cols']} columns | "
        f"Locations: {loc_str}. "
        + DOMAIN_SUMMARIES.get(primary, DOMAIN_SUMMARIES["general"])
    )

    sem_vals   = list(sem_map.values())[:15]
    domain_tags = [d for d,v in domain_scores.items() if v > 0]

    index_text = (
        f"File: {filepath.name} | Sheet: {sheet_name} | "
        f"Domain: {primary.replace('_',' ')} | "
        f"Rows: {profile['total_rows']} | Cols: {profile['total_cols']} | "
        f"Topics: {', '.join(domain_tags)} | "
        f"Contains: {', '.join(sem_vals[:12])} | "
        f"{summary[:350]}"
    )

    return {
        "doc_id":           hashlib.md5(f"{filepath.name}::{sheet_name}".encode()).hexdigest()[:12],
        "filename":         filepath.name,
        "filepath":         str(filepath),
        "file_type":        "xlsx",
        "sheet_name":       sheet_name,
        "title":            f"{primary.replace('_',' ').title()} — {sheet_name}",
        "inferred_domain":  primary,
        "domain_scores":    domain_scores,
        "col_semantic_map": sem_map,
        "flat_headers":     flat_hdrs,
        "total_rows":       profile["total_rows"],
        "total_cols":       profile["total_cols"],
        "data_start_row":   profile.get("data_start_row", 2),
        "summary":          summary,
        "index_text":       index_text,
        "product":          "Honeywell BA Operational Data",
        "product_version":  "live",
        "status":           "current",
        "language":         "en",
        "fetch_strategy":   "excel_rows",
        "fetch_params": {
            "filepath":         str(filepath),
            "sheet_name":       sheet_name,
            "data_start_row":   profile.get("data_start_row", 2),
            "flat_headers":     flat_hdrs,
            "col_semantic_map": sem_map,
        },
        "indexed_at": datetime.now().isoformat(),
    }


def build(rebuild: bool = False) -> dict:
    ensure_dirs()

    if UNIFIED_STORE.exists() and not rebuild:
        with open(UNIFIED_STORE,"rb") as f:
            store = pickle.load(f)
        print(f"Loaded existing unified store: {len(store['records'])} records")
        return store

    # Load PDF index
    pdf_records = []
    if PDF_INDEX.exists():
        with open(PDF_INDEX,"rb") as f:
            pdf_idx = pickle.load(f)
        pdf_records = pdf_idx.get("records",[])
        print(f"Loaded {len(pdf_records)} PDF records")
    else:
        print(f"Warning: no PDF index found at {PDF_INDEX}")
        print(f"Run: python3 src/ingestion/build_pdf_index.py first")

    # Extract Excel metadata
    excel_records = []
    xlsx_files = sorted(EXCEL_DIR.glob("*.xlsx"))
    print(f"\nExtracting metadata from {len(xlsx_files)} Excel files...")

    for xlsx in xlsx_files:
        try:
            wb = openpyxl.load_workbook(xlsx, read_only=True)
            sheets = wb.sheetnames
            wb.close()
        except Exception:
            continue
        print(f"\n  [{xlsx.name}]")
        for sheet in sheets:
            rec = build_excel_record(xlsx, sheet)
            if rec:
                excel_records.append(rec)
                print(f"    {sheet}: {rec['inferred_domain']} | "
                      f"{rec['total_rows']} rows | {rec['total_cols']} cols")

    all_records  = pdf_records + excel_records
    index_texts  = [r.get("index_text", r.get("summary","")) for r in all_records]

    vec = TfidfVectorizer(ngram_range=(1,2), max_features=8000, sublinear_tf=True)
    mat = vec.fit_transform(index_texts)

    store = {
        "records":      all_records,
        "vec":          vec,
        "mat":          mat,
        "pdf_count":    len(pdf_records),
        "excel_count":  len(excel_records),
        "built_at":     datetime.now().isoformat(),
    }
    with open(UNIFIED_STORE,"wb") as f:
        pickle.dump(store, f)

    print(f"\nUnified store saved: {UNIFIED_STORE.name}")
    print(f"  {len(all_records)} total records "
          f"({len(pdf_records)} PDF + {len(excel_records)} Excel)")
    print(f"  Matrix: {mat.shape} | {mat.data.nbytes/1024:.0f} KB")
    return store


if __name__ == "__main__":
    rebuild = "--rebuild" in sys.argv
    build(rebuild=rebuild)
