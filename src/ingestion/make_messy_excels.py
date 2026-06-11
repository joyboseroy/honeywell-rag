"""
ingestion/make_messy_excels.py
Generates three realistic messy enterprise Excel files into data/excel/.
Files have 57-72 cryptic columns, merged headers, multiple sheets.
Filenames give no clue to content — the metadata pipeline must discover domain.

Run from project root:
  python3 src/ingestion/make_messy_excels.py
"""

import sys, os
sys.path.insert(0, os.path.join(os.path.dirname(__file__), ".."))

import random
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime, timedelta
from pathlib import Path
from config import EXCEL_DIR, ensure_dirs

random.seed(99)

BUILDINGS = ["ATL-HQ","BLR-01","BLR-02","SNG-TW","UK-BRC"]
FLOORS    = ["B1","G","01","02","03","04","05"]
ZONES     = ["NE","NW","SE","SW","CORE","SRV","LIFT"]

def rdate(days_ago=400):
    return (datetime.now() - timedelta(days=random.randint(0, days_ago))).strftime("%d/%m/%Y")
def rint(lo, hi): return random.randint(lo, hi)
def rfloat(lo, hi, dp=2): return round(random.uniform(lo, hi), dp)
def rchoice(lst): return random.choice(lst)

thin = Side(style="thin", color="BBBBBB")
thin_border = Border(left=thin, right=thin, top=thin, bottom=thin)

def hdr(ws, row, col, val, bg="1F3864", fg="FFFFFF", sz=9):
    c = ws.cell(row=row, column=col, value=val)
    c.font  = Font(name="Calibri", size=sz, bold=True, color=fg)
    c.fill  = PatternFill("solid", fgColor=bg)
    c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    c.border = thin_border

def data(ws, row, col, val, bg=None):
    c = ws.cell(row=row, column=col, value=val)
    c.font   = Font(name="Calibri", size=8)
    c.border = thin_border
    if bg: c.fill = PatternFill("solid", fgColor=bg)


# ── FILE 1: Equipment field report — 68 columns ───────────────────────────────

def make_file1():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Fld_Rpt_Q3"

    SECTIONS = [
        ("ID/Loc",1,6),("Equip_Specs",7,18),("Inst_Params",19,30),
        ("Svc_Hist_Rcd",31,44),("Perf_KPI_Curr",45,56),
        ("Fin_Val_Calc",57,63),("Rmks_Oth",64,68),
    ]
    for label, cs, ce in SECTIONS:
        ws.merge_cells(start_row=1, start_column=cs, end_row=1, end_column=ce)
        hdr(ws, 1, cs, label, bg="C00000", sz=8)

    COLS = [
        "Rec_ID","Fac_Cd","Bld_Ref","Flr_Cd","Zn_Cd","Rm_No",
        "Eq_Cat","Eq_SubCat","Mfr_Cd","Mdl_Cd","Srl_No","Fw_Ver",
        "Hw_Rev","Cap_Val","Cap_Unt","Inp_V","Out_V","Prot_Cls",
        "Inst_Dt","Comm_Dt","Wrnt_Exp","Cfg_Ref","Ip_Addr","Mac_Addr",
        "Net_Seg","Baud_Rt","Comm_Prt","Poll_Iv","Wbus_Ad","Scl_Grp",
        "Lst_PM_Dt","Lst_PM_By","Lst_PM_Typ","Nxt_PM_Due","PM_Intv_Mo",
        "Lst_CM_Dt","CM_Typ_Cd","CM_Hrs","CM_Cst_LC","CM_Cst_FX",
        "Svc_Prvdr","WO_Ref","Flt_Cnt_12M","Dwntime_Hr",
        "Cur_Tmp_C","Cur_Hum_Pct","Cur_CO2_Ppm","Cur_Pwr_Kw","Enrg_Idx",
        "Alm_Cnt_7D","Alm_Cnt_30D","Thr_Viol_7D","Avail_Pct","Resp_Ms",
        "Lst_Rdg_Ts","Rdg_Qual",
        "Purch_Cst","Bk_Val_Cur","Dep_Rt_Pa","Dep_Mth","CAPEX_Yr",
        "Opex_Ann","LCOE_Calc",
        "Crit_Flg","Esc_Grp","Insp_Sts","Doc_Ref","Rmk_Fr",
    ]
    assert len(COLS) == 68
    for ci, h in enumerate(COLS, 1):
        hdr(ws, 2, ci, h, bg="2E4057", sz=8)
        ws.column_dimensions[get_column_letter(ci)].width = 9

    MFRS  = ["HW","TRN","SIE","JCI","ABB"]
    CATS  = ["AHU","FCU","VAV","CHR","BLR","FPL","IQ4","TC3","WMBS","AFPX"]
    PROTOS= ["BACnet","Modbus","LonWorks","DALI","Wallbus","Proprietary"]
    SVCS  = ["Inhouse","HW_SvcCo","ExtCo_A","ExtCo_B"]
    DEPS  = ["SL","DB","UoP"]
    INSP  = ["PASS","FAIL","PEND","OVRD","N/A"]

    for r in range(3, 153):
        bld   = rchoice(BUILDINGS)
        eq    = rchoice(CATS)
        mfr   = rchoice(MFRS)
        purch = rint(8000, 120000)
        age   = rfloat(0.5, 8)
        bkv   = max(0, int(purch * (1 - age * rfloat(0.08, 0.20))))
        alm7  = rint(0, 12)
        alm30 = rint(alm7, alm7+30)
        avail = rfloat(88, 99.9)
        crit  = 1 if (alm30 > 20 or avail < 92) else 0
        bg    = "FFF2CC" if crit else None

        vals = [
            f"R{r-2:05d}", f"FA-{rint(1000,9999)}", bld,
            rchoice(FLOORS), rchoice(ZONES), f"Rm-{rint(100,599)}",
            eq, f"{eq}_S{rint(1,4)}", mfr, f"{mfr}-{eq}-{rint(100,999)}",
            f"SN{rint(1000000,9999999)}", f"v{rint(2,5)}.{rint(0,9)}.{rint(0,9)}",
            f"HW-{rint(1,5)}", rfloat(0.5,100),
            rchoice(["kW","TR","VAC","kPa"]), rint(12,480), rint(12,240),
            rchoice(["IP20","IP54","IP65"]),
            rdate(2800), rdate(2700), rdate(365),
            f"CFG-{rint(1000,9999)}",
            f"10.{rint(1,254)}.{rint(1,254)}.{rint(1,254)}",
            f"AA:BB:{rint(10,99):02X}:{rint(10,99):02X}:{rint(10,99):02X}:{rint(10,99):02X}",
            f"VLAN{rint(10,99)}", rchoice([9600,19200,38400,115200]),
            rchoice(PROTOS), rint(5,60), rint(2,15), f"SG{rint(1,8)}",
            rdate(400), rchoice(SVCS), rchoice(["PM","INS","CAL"]),
            rdate(180), rint(3,24),
            rdate(600), rchoice(["EM","PM","CM","UPG","INS"]),
            rfloat(0.5,16), rfloat(200,8000), rfloat(200,8000),
            rchoice(SVCS), f"WO-{rint(10000,99999)}", alm30, rfloat(0,48),
            rfloat(16,30), rfloat(35,75), rint(350,1200),
            rfloat(0.5,85), rfloat(60,130),
            alm7, alm30, rint(0,alm7), avail, rint(50,500),
            datetime.now().strftime("%Y-%m-%d %H:%M"),
            rchoice(["GOOD","FAIR","POOR","UNK"]),
            purch, bkv, rfloat(0.08,0.20), rchoice(DEPS),
            rint(0, int(purch*0.2)), rfloat(1000,15000), rfloat(0.05,0.25),
            crit, f"ESC-{rint(1,5)}", rchoice(INSP),
            f"DOC-{rint(1000,9999)}",
            rchoice(["","Sensor drift","Comms fault","Filter overdue","Battery low",""]),
        ]
        assert len(vals) == 68
        for ci, v in enumerate(vals, 1):
            data(ws, r, ci, v, bg=bg)

    # Sheet 2 — building summary aggregate
    ws2 = wb.create_sheet("Bld_Smry_Agg")
    smry_cols = ["Bld_Ref","Tot_Eq","Act_Eq","Flt_Eq","PM_Ovrd",
                 "Alm_30D_Tot","Avg_Avail","OPEX_Ann_Tot","CAPEX_Yr_Tot","Crit_Cnt"]
    for ci, h in enumerate(smry_cols, 1):
        hdr(ws2, 1, ci, h, bg="1F3864")
    for bld in BUILDINGS:
        n = rint(20,60)
        ws2.append([bld, n, rint(int(n*0.8),n), rint(0,5), rint(0,8),
                    rint(10,200), rfloat(90,99.5),
                    rfloat(50000,500000), rfloat(20000,200000), rint(0,8)])

    # Sheet 3 — raw alarm log dump
    ws3 = wb.create_sheet("AlmLog_Raw")
    ws3.append(["TS","Eq_Ref","Alm_Cd","Alm_Txt","Sev","Ack_By","Ack_Ts","Dur_Min","Resol_Cd"])
    for _ in range(300):
        ts = datetime.now() - timedelta(minutes=rint(0,43200))
        ws3.append([
            ts.strftime("%Y-%m-%d %H:%M:%S"), f"R{rint(1,150):05d}",
            f"ALM-{rint(100,999)}",
            rchoice(["High temp","Comm fail","Sensor fault","CO2 high","Fan trip"]),
            rchoice(["P1","P2","P3","P4"]),
            rchoice(["Auto","J.Smith","R.Patel","M.Wong"]),
            (ts+timedelta(minutes=rint(1,120))).strftime("%Y-%m-%d %H:%M:%S"),
            rint(1,480), rchoice(["RST","ACK","ESC","OPN"]),
        ])

    fname = EXCEL_DIR / "Rpt_FA_v3_FINAL_USE_THIS.xlsx"
    wb.save(fname)
    print(f"  {fname.name} (68 cols, 3 sheets, 150 data rows)")


# ── FILE 2: Energy & utilities — 72 columns ───────────────────────────────────

def make_file2():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Mnth_Rpt"

    COLS = (
        ["Prd_Key","Yr","Mth_Nm","Mth_No","Wk_No","Qtr","Fac_Cd","Bld_Ref","Occ_Cls","Flr_Cnt"] +
        ["El_kWh_T","El_kWh_HVAC","El_kWh_Ltg","El_kWh_IT","El_kWh_Oth","El_PkDmd_kW",
         "El_OffPk_kWh","El_OnPk_kWh","El_PF","El_THD_Pct","El_Cost_LC"] +
        ["Gs_m3_T","Gs_kWh_Eq","Gs_Prs_mbar","Gs_Cost_LC","Gs_Efic_Pct"] +
        ["Wt_m3_T","Wt_m3_Cool","Wt_m3_San","Wt_m3_Irr","Wt_Cost_LC","Wt_Intns_m3pSqM"] +
        ["Th_Ctl_MWh","Th_Htg_MWh","Th_COP_Cool","Th_COP_Htg","Th_Setpt_C_D","Th_Setpt_C_N",
         "HVAC_Run_Hr","Chil_COP","Boil_Eff_Pct"] +
        ["CO2e_El_T","CO2e_Gs_T","CO2e_T_Tot","CO2e_Intns_kgSqM","CO2e_vs_Bsln_Pct",
         "GHG_Scpe1","GHG_Scpe2","GHG_Scpe3"] +
        ["Occ_Pct_Avg","Occ_Pk_Cnt","Flr_Area_Sqm","Wth_HDD","Wth_CDD","Wth_AvgT_C","Wth_RH_Pct"] +
        ["NRG_Cost_Tot_LC","NRG_Cost_USD","NRG_Bgt_USD","Bgt_Var_Pct","NRG_Intns_kWh_Sqm",
         "GRESB_Scr","EnStar_Rtg"] +
        ["Trgt_kWh","Trgt_Var_Pct","ISO50001_Cmp","LEED_Pts","Rpt_Sts","Verif_By","Rmk_NRG",
         "Aud_Flg","Aud_By"]
    )
    assert len(COLS) == 72

    SEC = [
        ("Prd_Fac_ID",1,10),("Electricity_Dtl",11,21),("Gas_Cons",22,26),
        ("Water_Cons",27,32),("Thrml_HVAC",33,41),("Carbon_GHG",42,49),
        ("Occ_Ctxt",50,56),("Fin_Bgt",57,63),("Tgt_Cmp",64,72),
    ]
    for label, cs, ce in SEC:
        ws.merge_cells(start_row=1, start_column=cs, end_row=1, end_column=ce)
        hdr(ws, 1, cs, label, bg="375623", sz=8)
    for ci, h in enumerate(COLS, 1):
        hdr(ws, 2, ci, h, bg="538135", sz=8)
        ws.column_dimensions[get_column_letter(ci)].width = 9

    MONTHS = ["Jan","Feb","Mar","Apr","May","Jun","Jul","Aug","Sep","Oct","Nov","Dec"]
    ri = 3
    for bld in BUILDINGS:
        for yr in [2022,2023,2024,2025]:
            for mi, mth in enumerate(MONTHS,1):
                if yr == 2025 and mi > 6: break
                el  = rint(40000,150000)
                gs  = rint(500,5000)
                wt  = rint(200,2000)
                co2 = el*0.000233 + gs*0.00202
                area= rint(3000,20000)
                cost= el*0.12 + gs*0.6 + wt*1.2
                bgt = cost * rfloat(0.85,1.15)
                trgt= el * rfloat(0.88,1.05)
                row = [
                    f"{bld}_{yr}_{mi:02d}", yr, mth, mi, rint(1,52), (mi-1)//3+1,
                    f"FA-{rint(1000,9999)}", bld, rchoice(["A","B","C"]), rint(3,10),
                    el, rint(int(el*0.4),int(el*0.55)), rint(int(el*0.1),int(el*0.2)),
                    rint(int(el*0.1),int(el*0.2)), rint(int(el*0.02),int(el*0.08)),
                    rint(200,600), rint(int(el*0.3),int(el*0.5)),
                    rint(int(el*0.4),int(el*0.6)), rfloat(0.85,0.99), rfloat(1,8),
                    round(cost*0.6,2),
                    gs, round(gs*10.5,1), rint(18,24), round(cost*0.15,2), rfloat(78,95),
                    wt, rint(int(wt*0.4),int(wt*0.7)), rint(int(wt*0.2),int(wt*0.4)),
                    rint(0,int(wt*0.1)), round(cost*0.05,2), round(wt/area,3),
                    rfloat(20,120), rfloat(5,60), rfloat(2.5,5.5), rfloat(2.8,4.2),
                    rfloat(20,24), rfloat(16,20), rint(400,720), rfloat(3,6), rfloat(80,95),
                    round(el*0.000233,2), round(gs*0.00202,2), round(co2,2),
                    round(co2*1000/area,2), rfloat(-20,15),
                    round(gs*0.00202,2), round(el*0.000233,2), rfloat(0,5),
                    rfloat(55,95), rint(50,500), area, rint(0,600), rint(0,400),
                    rfloat(-5,35), rfloat(40,80),
                    round(cost,2), round(cost/74,2), round(bgt/74,2),
                    round((cost-bgt)/bgt*100,1), round(el/area,2),
                    rint(30,90), rchoice(["Exc","Cert","None"]),
                    round(trgt,0), round((el-trgt)/trgt*100,1),
                    rchoice(["Y","N"]), rint(0,15),
                    rchoice(["FINAL","DRAFT","PEND"]),
                    rchoice(["R.Kumar","S.Patel","Auto","M.Lee"]), "",
                    rchoice(["Y","N",""]), rchoice(["Int","Ext",""]),
                ]
                assert len(row) == 72
                for ci, v in enumerate(row, 1):
                    ws.cell(row=ri, column=ci, value=v).font = Font(name="Calibri", size=8)
                ri += 1

    fname = EXCEL_DIR / "NRG_UTL_Cons_YTD_v2b.xlsx"
    wb.save(fname)
    print(f"  {fname.name} (72 cols, 2 sheets, {ri-3} data rows)")


# ── FILE 3: Commissioning records — 57 columns ────────────────────────────────

def make_file3():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Cx_Rec_All"

    COLS = (
        ["Cx_ID","Proj_No","Bld_Ref","Flr_Zn","Sys_Typ","Dev_Cat","Dev_Mfr",
         "Dev_Mdl","Dev_FW","Rev_No","Site_Ref"] +
        ["Inst_By","Cx_By","Cx_Dt","Cx_Sht_Ref","Clt_Rep","HW_Rep"] +
        ["IP_Addr","Subnet","GW_Addr","DNS1","DNS2","VLAN","Mac","Port_No","Proto"] +
        ["Wbus_Ad","SLC_No","SLC_Ad","Bac_Dev_Id","Bac_Net_No","Bac_Obj_Cnt"] +
        ["Sp_Heat_C","Sp_Cool_C","Sp_CO2_Ppm","Sp_Hum_Pct","Db_Heat","Db_Cool"] +
        ["DO_Qty","DI_Qty","AO_Qty","AI_Qty","UI_Qty",
         "Comm_Chk","DB_Chk","IO_Chk","Alm_Chk","Sch_Chk"] +
        ["Cx_Sts","Snag_No","Snag_Desc","Retst_Dt","Retst_By","Retst_Sts"] +
        ["Cert_No","Cert_Dt","Cert_Exp","Handover_Dt","Rmk_Cx"]
    )
    assert len(COLS) == 59

    for ci, h in enumerate(COLS, 1):
        hdr(ws, 1, ci, h, bg="7030A0", sz=8)
        ws.column_dimensions[get_column_letter(ci)].width = 10

    DEVS  = ["IQ4-Ctrl","TC300-Therm","RS-WMB","RD-WMB","AFP-3030",
             "FCM-1","ACM-48","VAV-Ctrl","FCU-Ctrl"]
    PROTO = ["BACnet/IP","BACnet/MSTP","Modbus-RTU","Wallbus","Proprietary","DALI"]
    STS   = ["PASS","FAIL","PEND","PART","HOLD"]
    MFRS  = ["HW","TRN","SIE","JCI"]

    for r in range(2, 202):
        dev  = rchoice(DEVS)
        sts  = rchoice(STS)
        snag = rint(0,3) if sts in ("FAIL","PART","HOLD") else 0
        bg   = "FFE0E0" if sts=="FAIL" else ("FFFACD" if sts in ("PEND","PART","HOLD") else None)

        vals = [
            f"CX-{r-1:05d}", f"PRJ-{rint(1000,9999)}", rchoice(BUILDINGS),
            f"{rchoice(FLOORS)}-{rchoice(ZONES)}",
            rchoice(["BEMS","Fire","Sec","HVAC","Ltg"]),
            dev, rchoice(MFRS), f"{dev[:3]}-{rint(100,999)}",
            f"v{rint(2,5)}.{rint(0,9)}", f"R{rint(1,5)}", f"SITE-{rint(100,999)}",
            f"Tech_{rchoice(['AJ','RK','SP','ML'])}",
            f"CxEng_{rchoice(['01','02','03'])}",
            rdate(400), f"CX-SHT-{rint(100,999)}",
            rchoice(["Yes","No","N/A"]), rchoice(["Present","Remote","N/A"]),
            f"10.{rint(1,254)}.{rint(1,254)}.{rint(1,254)}",
            "255.255.255.0", f"10.{rint(1,254)}.{rint(1,254)}.1",
            f"10.{rint(1,9)}.0.1", "8.8.8.8", f"VLAN{rint(10,99)}",
            f"AA:{rint(10,99):02X}:{rint(10,99):02X}:{rint(10,99):02X}:{rint(10,99):02X}:{rint(10,99):02X}",
            rint(1,65535), rchoice(PROTO),
            rint(2,15), rint(1,10), rint(1,127),
            rint(1000,65000), rint(1,50), rint(10,500),
            rfloat(18,24), rfloat(22,28), rint(700,1200),
            rfloat(40,65), rfloat(0.5,3), rfloat(0.5,3),
            rint(0,8), rint(0,12), rint(0,4), rint(0,16), rint(0,8),
            rchoice(["PASS","FAIL","N/A"]), rchoice(["PASS","FAIL","N/A"]),
            rchoice(["PASS","FAIL","N/A"]), rchoice(["PASS","FAIL","N/A"]),
            rchoice(["PASS","FAIL","N/A"]),
            sts, snag,
            f"SNAG-{rint(100,999)}: {rchoice(['Comm timeout','Wrong addr','IO mismatch','FW mismatch'])}" if snag else "",
            rdate(90) if sts=="FAIL" else "",
            f"CxEng_{rchoice(['01','02','03'])}" if sts=="FAIL" else "",
            rchoice(["PASS","PEND",""]) if sts=="FAIL" else "",
            f"CERT-{rint(10000,99999)}", rdate(200), rdate(365), rdate(100),
            rchoice(["","Cabling rerouted","Address conflict resolved","FW updated on site",""]),
        ]
        assert len(vals) == 59
        for ci, v in enumerate(vals, 1):
            data(ws, r, ci, v, bg=bg)

    fname = EXCEL_DIR / "Cx_TR_IQ4_WMB_AFP_SITE_Consolidated.xlsx"
    wb.save(fname)
    print(f"  {fname.name} (57 cols, 200 data rows)")


if __name__ == "__main__":
    ensure_dirs()
    print("Generating messy enterprise Excel files...")
    make_file1()
    make_file2()
    make_file3()
    print("Done. All files in data/excel/")
