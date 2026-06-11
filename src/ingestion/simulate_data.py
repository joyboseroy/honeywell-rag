"""
ingestion/simulate_data.py
Generates:
  data/sensors/honeywell_sensors.db   — 120 sensors, 80,760 readings
  data/excel/honeywell_asset_register.xlsx
  data/excel/honeywell_maintenance_log.xlsx
  data/excel/honeywell_energy_report.xlsx

Run from project root:
  python3 src/ingestion/simulate_data.py
"""

import sys
import os
sys.path.insert(0, os.path.join(os.path.dirname(__file__), ".."))

import sqlite3
import random
import openpyxl
from openpyxl.styles import Font, PatternFill
from datetime import datetime, timedelta
from pathlib import Path
from config import SENSOR_DIR, EXCEL_DIR, DB_PATH, ensure_dirs

random.seed(42)

BUILDINGS = [
    "Honeywell HQ Atlanta",
    "Honeywell Bangalore Campus",
    "Honeywell Singapore Tower",
    "Honeywell UK Bracknell",
]
FLOORS    = ["Floor 1","Floor 2","Floor 3","Floor 4","Floor 5"]
ZONES     = ["Zone A","Zone B","Zone C","Zone D"]
EQ_TYPES  = ["AHU","FCU","Chiller","Boiler","IQ4 Controller","AFP-3030","TC300","VAV"]
TECHS     = ["A.Kumar","S.Patel","R.Jones","M.Wong","J.Smith"]

def rdate(days_ago_max=730):
    return (datetime.now() - timedelta(days=random.randint(0, days_ago_max))).strftime("%Y-%m-%d")

def make_sensor_db():
    ensure_dirs()
    DB_PATH.unlink(missing_ok=True)
    conn = sqlite3.connect(str(DB_PATH))
    cur  = conn.cursor()

    cur.execute("""
        CREATE TABLE sensors (
            sensor_id       TEXT PRIMARY KEY,
            building        TEXT,
            floor           TEXT,
            zone            TEXT,
            sensor_type     TEXT,
            unit            TEXT,
            asset_id        TEXT,
            firmware_version TEXT
        )
    """)
    cur.execute("""
        CREATE TABLE readings (
            id          INTEGER PRIMARY KEY AUTOINCREMENT,
            sensor_id   TEXT,
            timestamp   TEXT,
            value       REAL,
            quality     TEXT,
            alert_flag  INTEGER DEFAULT 0
        )
    """)
    cur.execute("CREATE INDEX idx_sensor_ts ON readings(sensor_id, timestamp)")

    SENSOR_TYPES = {
        "TEMP":  ("°C",   18.0,  5.0,  26.0),
        "CO2":   ("ppm",  600.0, 200.0, 1000.0),
        "HUM":   ("%RH",  50.0,  15.0,  80.0),
        "OCC":   ("count",20.0,  20.0,  None),
        "HVAC":  ("kW",   15.0,  10.0,  None),
        "FIRE":  ("binary",0.0,  None,  0.5),
    }

    sensor_ids = []
    sid = 1
    for bld in BUILDINGS:
        for flr in FLOORS:
            for stype in SENSOR_TYPES:
                sensor_id = f"S{sid:04d}"
                unit, _, _, _ = SENSOR_TYPES[stype]
                cur.execute(
                    "INSERT INTO sensors VALUES (?,?,?,?,?,?,?,?)",
                    (sensor_id, bld, flr, random.choice(ZONES),
                     stype, unit, f"ASSET-{random.randint(1,200):04d}", f"v{random.randint(2,5)}.{random.randint(0,9)}")
                )
                sensor_ids.append((sensor_id, stype, *SENSOR_TYPES[stype]))
                sid += 1

    # 7 days of 15-min readings
    readings = []
    base = datetime.now() - timedelta(days=7)
    intervals = 7 * 24 * 4  # 672 readings per sensor

    for sensor_id, stype, unit, base_val, noise, alert_threshold in sensor_ids:
        for i in range(intervals):
            ts   = base + timedelta(minutes=i*15)
            hour = ts.hour
            # Realistic patterns
            if stype == "TEMP":
                val = base_val + random.gauss(0, noise * 0.3)
                if hour < 7 or hour > 20: val -= 2.0
            elif stype == "CO2":
                occ = 0.8 if 8 <= hour <= 18 else 0.1
                val = base_val + occ * random.gauss(0, noise) + random.gauss(0, 30)
            elif stype == "HUM":
                val = base_val + random.gauss(0, noise * 0.2)
            elif stype == "OCC":
                val = max(0, base_val * (0.9 if 9 <= hour <= 17 else 0.05) + random.gauss(0, 5))
            elif stype == "HVAC":
                val = max(0, base_val * (1.2 if 8 <= hour <= 18 else 0.3) + random.gauss(0, 2))
            elif stype == "FIRE":
                val = 1.0 if random.random() < 0.0001 else 0.0

            val  = round(max(0, val), 2)
            flag = 0
            if alert_threshold is not None and val > alert_threshold:
                flag = 1

            quality = "GOOD" if random.random() > 0.02 else "FAIR"
            readings.append((sensor_id, ts.strftime("%Y-%m-%d %H:%M:%S"), val, quality, flag))

    cur.executemany(
        "INSERT INTO readings (sensor_id,timestamp,value,quality,alert_flag) VALUES (?,?,?,?,?)",
        readings
    )
    conn.commit()
    conn.close()

    total = len(sensor_ids)
    print(f"  Sensor DB:      {DB_PATH.name}")
    print(f"  Sensors:        {total}")
    print(f"  Readings:       {len(readings):,}")


def hdr_cell(ws, row, col, val, bg="1F3864", fg="FFFFFF"):
    c = ws.cell(row=row, column=col, value=val)
    c.font  = Font(name="Calibri", size=9, bold=True, color=fg)
    c.fill  = PatternFill("solid", fgColor=bg)


def make_asset_register():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Asset Register"
    headers = ["Asset ID","Building","Floor","Zone","Equipment Type","Model",
               "Serial Number","Vendor","Install Date","Last Service Date",
               "Next Service Due","Warranty Expiry","Status","Assigned Tech",
               "Purchase Cost (USD)","Book Value (USD)"]
    for ci, h in enumerate(headers, 1):
        hdr_cell(ws, 1, ci, h)

    STATUSES = ["Operational","Under Maintenance","Fault","Decommissioned"]
    for i in range(1, 201):
        eq   = random.choice(EQ_TYPES)
        bld  = random.choice(BUILDINGS)
        cost = random.randint(5000, 80000)
        ws.append([
            f"ASSET-{i:04d}", bld, random.choice(FLOORS), random.choice(ZONES),
            eq, f"{eq[:3]}-{random.randint(100,999)}", f"SN{random.randint(1000000,9999999)}",
            random.choice(["Honeywell","Trend","Siemens","JCI"]),
            rdate(2500), rdate(400), rdate(180), rdate(365),
            random.choice(STATUSES), random.choice(TECHS),
            cost, int(cost * random.uniform(0.3, 0.9)),
        ])

    path = EXCEL_DIR / "honeywell_asset_register.xlsx"
    wb.save(path)
    print(f"  Asset register: {path.name} (200 assets)")


def make_maintenance_log():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Maintenance Log"
    headers = ["Log ID","Asset ID","Building","Equipment Type","Date",
               "Technician","Work Type","Description","Duration (hrs)",
               "Parts Used","Parts Cost (USD)","Labour Cost (USD)",
               "Total Cost (USD)","Result","Next Action"]
    for ci, h in enumerate(headers, 1):
        hdr_cell(ws, 1, ci, h)

    WORK_TYPES = ["Preventive Maintenance","Corrective","Emergency Repair",
                  "Inspection","Calibration","Upgrade"]
    RESULTS    = ["Completed","Deferred","Escalated","Completed - Parts Required"]
    for i in range(1, 501):
        eq    = random.choice(EQ_TYPES)
        parts = random.randint(0, 2000)
        labour= random.randint(200, 3000)
        ws.append([
            f"LOG-{i:05d}", f"ASSET-{random.randint(1,200):04d}",
            random.choice(BUILDINGS), eq, rdate(700),
            random.choice(TECHS), random.choice(WORK_TYPES),
            f"{eq} {random.choice(WORK_TYPES).lower()}",
            round(random.uniform(0.5, 12), 1),
            f"Filter set, gasket" if random.random()>0.5 else "",
            parts, labour, parts+labour,
            random.choice(RESULTS),
            "Schedule follow-up" if random.random()>0.7 else "",
        ])

    path = EXCEL_DIR / "honeywell_maintenance_log.xlsx"
    wb.save(path)
    print(f"  Maintenance log: {path.name} (500 entries)")


def make_energy_report():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Monthly Energy Report"
    headers = ["Month","Building","HVAC kWh","Lighting kWh","IT Load kWh",
               "Total kWh","Carbon kg CO2","Cost (USD)","Target kWh",
               "Variance %","Peak Demand kW","Occupancy %"]
    for ci, h in enumerate(headers, 1):
        hdr_cell(ws, 1, ci, h)

    months = ["Jan","Feb","Mar","Apr","May","Jun",
              "Jul","Aug","Sep","Oct","Nov","Dec"]
    for yr in [2024, 2025]:
        for mi, mth in enumerate(months, 1):
            if yr == 2025 and mi > 6: break
            for bld in BUILDINGS:
                total  = random.randint(40000, 120000)
                target = int(total * random.uniform(0.9, 1.1))
                ws.append([
                    f"{mth} {yr}", bld,
                    int(total*0.45), int(total*0.15), int(total*0.20),
                    total, round(total * 0.000233, 1),
                    round(total * 0.12, 2), target,
                    round((total-target)/target*100, 1),
                    random.randint(200,500),
                    round(random.uniform(55, 95), 1),
                ])

    path = EXCEL_DIR / "honeywell_energy_report.xlsx"
    wb.save(path)
    print(f"  Energy report:  {path.name} (96 records)")


if __name__ == "__main__":
    ensure_dirs()
    print("Generating Honeywell BA simulation data...")
    make_sensor_db()
    make_asset_register()
    make_maintenance_log()
    make_energy_report()
    print("Done. All files written to data/")
