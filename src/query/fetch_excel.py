"""
query/fetch_excel.py
Fetches relevant rows from Excel files using semantic column labels.
Columns are already mapped to plain English in the metadata record.
Falls back to sample rows if no keyword match found.
"""

import sys, os
sys.path.insert(0, os.path.join(os.path.dirname(__file__), ".."))

import openpyxl
from pathlib import Path
from config import MAX_EXCEL_ROWS


def fetch(record: dict, query: str) -> dict:
    """
    Fetch rows from the original Excel file that match the query.
    Uses semantic column labels from metadata — not cryptic column names.
    """
    params    = record.get("fetch_params", {})
    filepath  = params.get("filepath", record.get("filepath",""))
    sheet     = params.get("sheet_name", record.get("sheet_name",""))
    sem_map   = params.get("col_semantic_map", {})
    flat_hdrs = params.get("flat_headers", [])
    start     = params.get("data_start_row", 2)

    source = f"{record.get('filename','')} :: {sheet}"

    if not filepath or not Path(filepath).exists():
        return {
            "lane":   "Excel",
            "source": source,
            "score":  record.get("relevance_score",0),
            "text":   f"[File not found: {filepath}]",
        }

    try:
        wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
    except Exception as e:
        return {
            "lane":   "Excel",
            "source": source,
            "score":  record.get("relevance_score",0),
            "text":   f"[Cannot open file: {e}]",
        }

    ws = wb[sheet] if sheet and sheet in wb.sheetnames else wb.active
    q_words     = set(query.lower().split())
    rows        = []
    all_rows_read = []

    # Also expand query words with synonyms for fault/alarm
    expanded_words = set(q_words)
    synonym_map = {
        "fault": ["alm","flt","alarm","error","fail"],
        "alarm": ["alm","flt","fault","alert"],
        "event": ["alm","flt","cnt","count"],
        "most":  [],
    }
    for w in list(q_words):
        for syn in synonym_map.get(w, []):
            expanded_words.add(syn)

    for row in ws.iter_rows(min_row=start, values_only=True):
        row_str = " ".join(str(v or "").lower() for v in row)
        if any(w in row_str for w in expanded_words if len(w) > 2):
            row_data = _extract_row(row, flat_hdrs, sem_map, max_cols=14)
            if row_data:
                rows.append(row_data)
        all_rows_read.append(row)
        if len(rows) >= MAX_EXCEL_ROWS:
            break
    wb.close()

    if rows:
        content = _format_rows(source, rows[:10], "matching rows")
    else:
        # Fallback: if sheet is small (<=20 rows) send everything — always useful to LLM
        # If sheet is large, send first 8 rows with semantic labels
        limit = len(all_rows_read) if record.get("total_rows",999) <= 20 else 8
        fallback = []
        for row in all_rows_read[:limit]:
            rd = _extract_row(row, flat_hdrs, sem_map, max_cols=14)
            if rd:
                fallback.append(rd)
        if fallback:
            content = _format_rows(source, fallback, f"all {len(fallback)} rows" if limit <= 20 else "sample rows")
        else:
            content = (
                f"No data found in {source}. "
                f"Sheet contains {record.get('total_rows',0)} rows of "
                f"{record.get('inferred_domain','').replace('_',' ')} data."
            )

    return {
        "lane":    "Excel",
        "source":  source,
        "version": "live",
        "score":   record.get("relevance_score",0),
        "text":    content,
    }


def _extract_row(row: tuple, flat_hdrs: list, sem_map: dict, max_cols: int = 14) -> dict:
    """Extract a row dict using semantic column labels."""
    rd = {}
    for ci, val in enumerate(row[:max_cols]):
        if val is not None:
            col_name = flat_hdrs[ci] if ci < len(flat_hdrs) else f"Col{ci}"
            label    = sem_map.get(col_name, col_name)
            rd[label[:45]] = str(val)[:45]
    return rd


def _format_rows(source: str, rows: list, label: str) -> str:
    lines = [f"{source} ({label}):"]
    for i, r in enumerate(rows, 1):
        row_str = " | ".join(
            f"{k}: {v}" for k, v in r.items()
            if v and str(v).strip()
        )
        lines.append(f"  [{i}] {row_str}")
    return "\n".join(lines)
